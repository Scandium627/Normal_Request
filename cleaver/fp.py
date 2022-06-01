# -*- coding:utf-8 -*-
# author: Scandium
# work_location: Bei Jing
# File : fp.py
# time: 2021/4/27 16:20


import datetime
import random
import time

from openpyxl.reader.excel import load_workbook


# work_place = "科电大厦"
# list_place = [["复兴路", 2, ["开发联调", "部署服务", "现场维护"]]]


def build_dic_invoices(list_place_input):
    dic_invoices = {}
    for place in list_place_input:
        if place[1] not in dic_invoices.keys():
            dic_invoices[place[1]] = [{"place": place[0], "affairs": place[2]}]
        else:
            dic_invoices[place[1]].append({"place": place[0], "affairs": place[2]})
    return dic_invoices

#
# input_list_test = [["2021-04-27", 27], ["2021-04-22", 14]]


def distance_to_info(distance_input, dic_invoices):
    if distance_input in dic_invoices:
        info = random.choice(dic_invoices[distance_input])
        info_place = info['place']
        info_affairs = random.choice(info['affairs'])
        info_out = [info_place, info_affairs]
    else:
        info_out = ["", ""]
    return info_out


# 日期转换星期
def get_week_day(date):
    week_day_dict = {
        0: '星期一',
        1: '星期二',
        2: '星期三',
        3: '星期四',
        4: '星期五',
        5: '星期六',
        6: '星期天',
    }
    day = datetime.datetime.strptime(date, "%Y-%m-%d").weekday()
    return week_day_dict[day]


# get_week_day("2021-04-27")
#
# input_list_test = [["2021-04-27", 27], ["2021-04-22", 14]]


def out_invoices(input_list, dic_invoices,home ="科电"):
    output_dic = {}
    for invoice in input_list:
        distance = int(invoice[1] / 10)
        if invoice[0] not in output_dic.keys():  # [date，route（home,place）,cost,affair]
            date = invoice[0]
            cost = invoice[1]
            weekday = get_week_day(date)
            [place, affair] = distance_to_info(distance, dic_invoices)
            route = f"{home}至{place}"
            output_dic[invoice[0]] = [[date, weekday, route, cost, affair]]
            # print(output_dic)
        elif invoice[0] in output_dic.keys() and len(output_dic[invoice[0]]) % 2 == 1:
            last_info = output_dic[invoice[0]][-1]
            # print(last_info[2])
            place = last_info[2].split("至")[1]
            #home = last_info[2].split("至")[1]
            route = f"{place}至{home}"
            date = invoice[0]
            cost = invoice[1]
            affair = "回所"
            weekday = get_week_day(date)
            output_dic[invoice[0]].append([date, weekday, route, cost, affair])
            # print(output_dic)
        elif invoice[0] in output_dic.keys() and len(output_dic[invoice[0]]) % 2 == 0:
            date = invoice[0]
            cost = invoice[1]
            weekday = get_week_day(date)
            [place, affair] = distance_to_info(distance, dic_invoices)
            route = f"{home}至{place}"
            output_dic[invoice[0]].append([date, weekday, route, cost, affair])
    return output_dic


# a11 = out_invoices(input_list_test,dic_invoices)


def dic_to_list(input_dic):
    out_list = []
    key_list = sorted(list(input_dic.keys()))
    for key in key_list:
        out_list.extend(input_dic[key])
    return out_list


# 表格操作

# 分割列表数据每num个一组
def data_depart(input_list, depart_by_num):
    if input_list:
        out_list = []
        num = len(input_list)
        if num % depart_by_num == 0:
            sheets_num = int(num / depart_by_num)
        else:
            sheets_num = int(num / depart_by_num) + 1
        for i in range(sheets_num):
            try:
                out_list.append(input_list[depart_by_num * i:depart_by_num * (i + 1)])
            except:
                out_list.append(input_list[depart_by_num * i:])
        return out_list
    else:
        return []


def total_cost_cal(input_list):
    total_cost = 0
    sum_page_dic = {}
    for page_num in range(len(input_list)):
        sum_now = sum([int(invoice[3]) for invoice in input_list[page_num]])
        total_cost += sum_now
        sum_page_dic[page_num] = sum_now
    return total_cost, sum_page_dic


def out_save_file(input_invoices_list, mobile_file_path):  # [date, weekday, route, cost, affair]
    # 加载模板并且打开相应sheet
    toal_cost, sum_page_dic = total_cost_cal(input_invoices_list)
    for page_num in range(len(input_invoices_list)):
        wb = load_workbook(filename=mobile_file_path)
        # sheetnames = wb.get_sheet_names()
        ws = wb["Sheet1"]
        page_cost_sum = sum_page_dic[page_num]
        for invoice_num in range(len(input_invoices_list[page_num])):
            invoice_data = input_invoices_list[page_num][invoice_num]
            month = invoice_data[0].split("-")[1]
            day = invoice_data[0].split("-")[2]
            day_month = f" {month} 月 {day} 日"
            ws.cell(row=invoice_num + 4, column=2, value=day_month)  # 写入日期
            ws.cell(row=invoice_num + 4, column=3, value=invoice_data[1])  # 写入星期
            ws.cell(row=invoice_num + 4, column=4, value=invoice_data[2])  # 写入路径
            ws.cell(row=invoice_num + 4, column=5, value=invoice_data[4])  # 写入事由
            ws.cell(row=invoice_num + 4, column=6, value=invoice_data[3])  # 写入金额
        ws.cell(row=24, column=9, value=page_cost_sum)  # 写入本业金额
        ws.cell(row=25, column=9, value=toal_cost)  # 写入总金额
        date_now = time.strftime("%Y-%m-%d", time.localtime())
        wb.save(f"invoice_{date_now}_{page_num}.xlsx")


def read_invoice_list(mobile_file_path):
    wb = load_workbook(filename=mobile_file_path)
    sheetnames = wb.get_sheet_names()
    print(sheetnames)
    ws = wb['Sheet1']
    rows = ws.max_row
    list_output = []
    for row in range(rows):
        list_output.append([str(ws.cell(row=row + 1, column=1).value).split(" ")[0], ws.cell(row=row + 1, column=2).value])
    return list_output

if __name__ == "main":
    work_place = "科电大厦"
    list_place = [["五道口", 1, ["打印文档", "文档传递"]],
                  ["计算所", 1, ["文档传递", "数据传递"]],
                  ["中关村", 1, ["打印文档", "数据传递"]],
                  ["中科雨辰", 2, ["开发联调", "部署服务", "现场维护"]],
                  ["军科", 3, ["开发联调", "部署服务", "现场维护"]],
                  ["未来城", 4, ["开发联调", "部署服务", "现场维护"]],
                  ["邓庄南路", 3, ["文件盖章", "文件传递", "文件签字"]],
                  ["28所", 5, ["文件传递", "光盘传递"]],
                  ["军科", 6, ["文件传递", "开发联调"]],
                  ["未来城", 7, ["服务部署", "开发联调"]],
                  ["未来城", 8, ["服务部署", "开发联调"]],
                  ["航天研究院", 9, ["文件传递", "光盘传递"]],
                  ["航天研究院", 10, ["文件传递", "光盘传递"]],
                  ["航天研究院", 11, ["文件传递", "光盘传递"]]]
    dic_invoices = build_dic_invoices(list_place)
    invoices_list = read_invoice_list("发票.xlsx")
    dic_data = out_invoices(invoices_list, dic_invoices)
    input_invoices_list = data_depart(dic_to_list(dic_data),20)
    out_save_file(input_invoices_list,"市内交通登记.xlsx")
    pass








# _______________此行之下______皆为测试
#
# list_Test = []
# for i in range(50):
#     list_Test.append(i)
#
# print(data_depart(list_Test, 9))
#
# input_invoices_list = [[["1l", 1], ["2l", 2]], [["1l", 44], ["1l", 3]]]
# total_cost = [i]
# total_c = []
# for row_num in range(0, 20):
#     print(row_num)
#     print(row_num + 4)
#
# total_cost_cal(input_invoices_list)
#
#
# class info_build:
#     def __int__(self):
#         self.place_list = {}
#         self.cost_list = {}
#         self.order = []
#         self.word_place = {}
#
#     def place_in(self, place_input):  # format = [{北京:[0,1]}]
#         place_combine = {}
#         for place in place_input:
#             for place_2 in place_input:
#                 the_cost = list(place.values())[0]
#
#
#
#
# wb = load_workbook(filename=r'市内交通登记.xlsx')
# sheetnames = wb.get_sheet_names()
# ws = wb.get_sheet_by_name(sheetnames[0])
#
# print("Work Sheet Titile:", ws.title)
# print("Work Sheet Rows:", ws.max_row)
# print("Work Sheet Cols:", ws.max_column)
#
# w1 = ws.cell(row=2, column=2).value
#
# ws.cell(row=4, column=7, value="item")
# print(w1)
#
# wb.save('test_case.xlsx')